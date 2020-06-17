from openpyxl import Workbook, load_workbook


class Xl:
    def __init__(self):
        pass

    def load_input_data(self, file_path):
        wb = load_workbook(file_path)
        ws = wb[wb.sheetnames[0]]

        data = {}
        i = 2
        while True:
            num = ws.cell(row=i, column=1).value
            try:
                num = int(num)
            except (TypeError, ValueError):
                break

            data[num] = [
                float(ws.cell(row=i, column=2).value),
                float(ws.cell(row=i, column=3).value),
                int(ws.cell(row=i, column=4).value),
            ]

            i += 1

        return data

    def create_workbook(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(["Time", "Voltage(V)", "Current(A)"])

    def update_data(self, data):
        self.ws.append(data)

    def save_workbook(self, path):
        self.wb.save(path)
